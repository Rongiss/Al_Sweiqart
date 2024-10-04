import openpyxl
import os
import pprint
import copy


def main():
    wb = openpyxl.load_workbook('Шаблон ВОР (задание) [kaSta0] — копия.xlsx')
    all_list = dict()
    for i in wb.sheetnames:
        di = dict()
        sh = wb[i]
        for col in range(1, sh.max_column + 1):
            l = []
            for roww in range(1, sh.max_row + 1):
                l.append(sh.cell(row=roww, column=col).value)
            di[col] = copy.deepcopy(l)
            l.clear()
        all_list[i] = di
    pprint.pprint(len(all_list['Виды работ'][3]))
    pprint.pprint(len(all_list['Справочник'][3]))


def main2():
    wb = openpyxl.load_workbook('example.xlsx')
    print(wb.sheetnames)
    sheet = wb.active
    print(sheet.columns[1])


if __name__ == '__main__':
    # main()

    main2()
